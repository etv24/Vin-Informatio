
import requests,json
import openpyxl
import json
import os



line = 2
nextvinstringlength = 17
while line < 6 and nextvinstringlength == 17:

    vincell = 'A' + str(line)
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'Money.xlsx')
    sheet_ranges = wb['Vehicles']
    vin = sheet_ranges[vincell].value
    vinstr = str(vin)

    a,b,c,d,e,f,g,h,i,j,k,l,m,n,o,p,q = vinstr

    if j=='Y':
        mfgyear = 2000
    elif j == '1':
        mfgyear = 2001
    elif j == '2':
        mfgyear = 2002
    elif j == '3':
        mfgyear = 2003
    elif j == '4':
        mfgyear = 2004
    elif j == '5':
        mfgyear = 2005
    elif j == '6':
        mfgyear = 2006
    elif j == '7':
        mfgyear = 2007
    elif j == '8':
        mfgyear = 2008
    elif j == '9':
        mfgyear = 2009
    elif j == 'A':
        mfgyear = 2010
    elif j == 'B':
        mfgyear = 2011
    elif j == 'C':
        mfgyear = 2012
    elif j == 'D':
        mfgyear = 2013
    elif j == 'E':
        mfgyear = 2014
    elif j == 'F':
        mfgyear = 2015
    elif j == 'G':
        mfgyear = 2016
    elif j == 'H':
        mfgyear = 2017
    elif j == 'I':
        mfgyear = 2018
    elif j == 'J':
        mfgyear = 2019
    elif j == 'K':
        mfgyear = 2020

    print('working...')
    url = 'https://vpic.nhtsa.dot.gov/api/vehicles/decodevin/' + vinstr + '?format=json&modelyear=' + str(mfgyear);
    js = requests.get(url).json()
    make = js['Results'][6]['Value']
    model = js['Results'][8]['Value']
    year = js['Results'][9]['Value']


    makecell = 'B' + str(line)
    modelcell = 'C' + str(line)
    yearcell = 'D' + str(line)
    sheet_ranges[makecell].value = make
    sheet_ranges[modelcell].value = model
    sheet_ranges[yearcell].value = year

    print('l'+str(line)+' - Vin: '+vinstr+' - Make: '+str(make)+ ' - Model: '+str(model)+' - Year: '+str(mfgyear))

    wb.save('Money.xlsx')

    vincell = 'A' + str(line+1)
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'Money.xlsx')
    sheet_ranges = wb['Vehicles']
    vin = sheet_ranges[vincell].value
    vinstr = str(vin)

    nextvinstringlength = len(vinstr)
    line = line+1
    

print('Program Has Finished')










# from openpyxl import load_workbook
# wb = load_workbook(filename = 'Money.xlsx')
# sheet_ranges = wb['Vehicles']
# print(sheet_ranges['A2'].value)

# url = 'https://vpic.nhtsa.dot.gov/api/vehicles/decodevin/5UXWX7C5*BA?format=json&modelyear=2011';
# r = requests.get(url);
# '''rdata = r.json()'''
# '''print(r.text)'''

# js = requests.get(url).json()
# randdata = js['Results'][23]['Value']
# print(randdata)

# sheet_ranges['B2'].value = 6
# wb.save('Money.xlsx')
